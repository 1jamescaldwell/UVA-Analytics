# James Caldwell, Fall 2025

import pandas as pd
import numpy as np
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import Font
from dotenv import load_dotenv

def load_sheets_with_dynamic_header(file_path):
# This function loads all sheets from an Excel file, checking the first two rows for the header row.
# It skips sheets named 'summary', 'changelog', 'warnings', 'errors', and 'vcsin xref'.
# If a sheet cannot be read, it is skipped with a warning printed to the console.

    result = {}
    meta_data = {}
    excel_file = pd.ExcelFile(file_path)
    for sheet_name in excel_file.sheet_names:
        try:
            
            if sheet_name.strip().lower() not in {'summary', 'changelog', 'warnings', 'errors', 'vcsin xref','comparison summary'}:

                # result[sheet_name] = pd.read_excel(excel_file, sheet_name=sheet_name, header=2) # Default to header on row 3 (index 2)
                preview = excel_file.parse(sheet_name=sheet_name, header=2)
                result[sheet_name] = preview #.iloc[header_row + 1:].reset_index(drop=True)
                meta_data[sheet_name] = excel_file.parse(sheet_name=sheet_name, nrows=2)

        except Exception as e:
            print(f"Skipped sheet '{sheet_name}' due to error: {e}")
            continue

    return result, meta_data

def select_file(title):
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    file_path = filedialog.askopenfilename(
        title=f"Select {title} Excel file",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    return file_path

def collect_notes_columns(df,sheet_name):
    """
    This function checks if any column in the DataFrame contains 'comments' (case insensitive).
    If found, it returns a DataFrame with those columns.
    """

    comment_cols = df[[col for col in df.columns if 'comments' in col.lower()]]

    # first try to add SOCSEC1 and Rowid if they exist. Error count will be 2 if both are missing
    try:
        notes_df = comment_cols.copy()
        notes_df['SOCSEC1'] = df['SOCSEC1'].copy()
    except:
        print('Error finding comments for sheet: ' + sheet_name + 'col names: ' + str(df.columns))
    return notes_df if not notes_df.empty else None

def change_log(file1_path, file2_path, result_path):

    # Load all sheets, using header on row 2 (index 1)  
    xl1,xl1_meta = load_sheets_with_dynamic_header(file1_path)
    xl2,xl2_meta = load_sheets_with_dynamic_header(file2_path)

    file_name1 = file1_path.split('/')[-1][:-5] # Remove .xlsx and filepath
    file_name2 = file2_path.split('/')[-1][:-5]

    writer = pd.ExcelWriter(result_path, engine='openpyxl')
    summary = []#[f'V1 file: {file_name1}', f'V2 file: {file_name2}']

    pd.DataFrame({"Summary": summary}).to_excel(writer, sheet_name="Comparison Summary", index=False)

    # Get union of all sheet names
    all_sheet_names = set(xl1.keys()).union(set(xl2.keys()))
    all_sheet_names = sorted(all_sheet_names)  # Sort sheet names for consistent order

    for sheet in all_sheet_names:

        # If either sheet is missing from V1 or V2, df1 or df2 will be None
        df1 = xl1.get(sheet)
        df2 = xl2.get(sheet)

        # Drop FAKeyint column
        col_to_ignore = 'FAKeyint'
        if df1 is not None and col_to_ignore in df1.columns:
            df1 = df1.drop(columns=[col_to_ignore])
        if df2 is not None and col_to_ignore in df2.columns:
            df2 = df2.drop(columns=[col_to_ignore])

        if df1 is not None and df2 is not None:
            
            # We can only merge reliably on SOCSEC1, so if it's missing from either file OR there are duplicate rows for SOCSEC1, just copy over the latest file sheet as is and note in summary
            if 'SOCSEC1' not in df2.columns or ('SOCSEC1' in df2.columns and df2['SOCSEC1'].duplicated().any()):
                
                # Merge notes manually for these on SOCSEC1. 
                # Add any to this list if they have a print out of comments from the output
                # This is a bit of blind merge, but it seems to be working? We may have to merge on Socsec1 + Repyear or REPPER. 
                if sheet in ['BFE001W04','BNB001W03','BNB006W01','BNB006W05','BSI003W01']:
                    df2 = df2.drop_duplicates()
                    df1 = df1.drop_duplicates()
                    df1_comments_col = [col for col in df1.columns if 'comments' in col.lower()]
                    df2 = pd.merge(df2,df1[['SOCSEC1']+df1_comments_col], how='left', on=['SOCSEC1'])

                for col in df1.columns:
                    if 'comments' in col.lower():
                        print(f"Unique values in {sheet}:'{col}':")
                        print(df1[col].dropna().unique())
                
                df2.to_excel(writer, sheet_name=sheet, index=False,startrow=2) # Write the original DataFrame to the output file, since it still has the Notes columns
                summary.append(f"Error '{sheet}' present in latest file. No SSN column to compare or duplicate SSN rows, so no summary/comparison stats.")
                # Add back the two rows of metadata (Error description and link) from the first file
                ws = writer.sheets[sheet]
                ws.cell(row=1, column=1, value=xl2_meta[sheet].columns[0]) # Error description
                ws.cell(row=2, column=3, value=xl2_meta[sheet].iloc[0, 2]) # Link

            else:
                if df2['SOCSEC1'].duplicated().any():
                    print(sheet + " " + str(df2['SOCSEC1'].duplicated().sum()))
                    print(df2.head())

                # Rename NaN column names to "missing"
                df1.columns = ["missing col name" if pd.isna(col) else col for col in df1.columns]
                df2.columns = ["missing col name" if pd.isna(col) else col for col in df2.columns]
                

                # Standardize column order
                df1_sorted = df1[sorted(df1.columns)]
                df2_sorted = df2[sorted(df2.columns)]
                # Try to sort by SSID, then Rowid if SSID doesn't exist
                try:
                    df1_sorted = df1_sorted.sort_values(by='SSID') 
                    df2_sorted = df2_sorted.sort_values(by='SSID')
                except:
                    try:
                        df1_sorted = df1_sorted.sort_values(by='Rowid') 
                        df2_sorted = df2_sorted.sort_values(by='Rowid')
                    except:
                        print(sheet + ' has no SSID or Rowid column to sort by.')
                        pass

                # Remove 'Comments' columns if they exist, and collect them separately to be added back later after comparison
                if df1_sorted.columns.str.lower().str.contains('comments').any() or df2_sorted.columns.str.lower().str.contains('comments').any(): # Check if any column name contains 'notes' (Notes, notes, notes:, etc)

                    notes1 = collect_notes_columns(df1_sorted, sheet)
                    notes2 = collect_notes_columns(df2_sorted, sheet)

                    df1_sorted = df1_sorted.drop(columns=[col for col in df1_sorted.columns if 'comments' in col.lower()])
                    df2_sorted = df2_sorted.drop(columns=[col for col in df2_sorted.columns if 'comments' in col.lower()])

                if df1_sorted.equals(df2_sorted): # If files are identical
                    
                    # Add v2 comments
                    # df1 = df1 + notes2.columns[:-1] 
                    df1 = pd.concat([df1, notes2.iloc[:, :-1]], axis=1) # add all notes2 columns except SOCSEC1, which is last
                    df1['New Error?'] = 'No' # Add New Error? column with 'No' since there are no new errors
                    df1.to_excel(writer, sheet_name=sheet, index=False,startrow=2) # Write the original DataFrame to the output file, since it still has the Notes columns
                    summary.append(f"Error '{sheet}': Exists in both files, no differences.")
                    ws = writer.sheets[sheet]
                    ws.cell(row=1, column=1, value=xl2_meta[sheet].columns[0]) # Error description
                    ws.cell(row=2, column=3, value=xl2_meta[sheet].iloc[0, 2]) # Link
                else:
                    df1_tagged = df1_sorted.copy()
                    df1_tagged["__source__"] = "File 1"

                    df2_tagged = df2_sorted.copy()
                    df2_tagged["__source__"] = "File 2"

                    combined = pd.concat([df1_tagged, df2_tagged], ignore_index=True)
                    
                    combined["__duplicated__"] = combined.duplicated(subset=combined.columns.difference(["__source__"]), keep=False)

                    def classify(row):
                        if row["__duplicated__"]:
                            return "Both"
                        return row["__source__"]

                    combined["New Error?"] = combined.apply(classify, axis=1)
                    output_df = combined.drop(columns=["__source__", "__duplicated__"])

                    # Drop duplicate rows, which occurs if there are identical rows in both files. Only keep one instance.
                    output_df = output_df.drop_duplicates().reset_index(drop=True)

                    # Add notes columns back in if they exist
                    try:
                        if notes1 is not None:
                            output_df = output_df.merge(notes1, how='left', on='SOCSEC1')
                        if notes2 is not None:
                            output_df = output_df.merge(notes2, how='left', on='SOCSEC1')
                    except:
                        output_df['Notes'] = 'Error merging notes columns from python script. Check manually.'
                        print(f'Error merging notes columns in {sheet}.')

                    # We don't care anymore about errors from only file 1, drop them.
                    output_df = output_df[output_df["New Error?"] != "File 1"]

                    output_df["New Error?"] = output_df["New Error?"].replace({
                        "File 2": "Yes",
                        "Both": "No"
                    })

                    only_in_file1 = (combined["New Error?"] == "File 1").sum()
                    only_in_file2 = (combined["New Error?"] == "File 2").sum()
                    in_both_files = (combined["New Error?"] == "Both").sum()

                    # We don't care anymore about errors from only file 1, drop them.
                    combined = combined[combined["New Error?"] != "File 1"]

                    output_df.to_excel(writer, sheet_name=sheet, index=False,startrow=2)
                    summary.append(f"Error '{sheet}': "
                                    f"{only_in_file1} errors gone, {in_both_files} previous errors still present, {only_in_file2} new errors from '{file_name2}'")
                    # Add back the two rows of metadata (Error description and link) from the first file
                    ws = writer.sheets[sheet]
                    ws.cell(row=1, column=1, value=xl2_meta[sheet].columns[0]) # Error description
                    ws.cell(row=2, column=3, value=xl2_meta[sheet].iloc[0, 2]) # Link
        elif df1 is not None:
            summary.append(f"Error '{sheet}': no longer present in latest file.")
        elif df2 is not None:
            df2.to_excel(writer, sheet_name=sheet, index=False, startrow=2) # Write the original DataFrame to the output file, since it still has the Notes columns
            summary.append(f"Error '{sheet}' added to latest file.")
            # Add back the two rows of metadata (Error description and link) from the first file
            ws = writer.sheets[sheet]
            ws.cell(row=1, column=1, value=xl2_meta[sheet].columns[0]) # Error description
            ws.cell(row=2, column=3, value=xl2_meta[sheet].iloc[0, 2]) # Link

    # Write summary
    pd.DataFrame({"Summary": summary}).to_excel(writer, sheet_name="Comparison Summary", index=False)
    writer.close()
    
def classify_status(text):
    if 'longer' in str(text).lower():
        return 'Complete'
    elif 'added' in str(text).lower():
        return 'New Error'
    else:
        return 'Recurring error'

def summary_page(v1_file_path, result_path):
    summary_df_new =pd.read_excel(result_path, sheet_name='Comparison Summary')
    summary_df_old =pd.read_excel(v1_file_path, sheet_name='Comparison Summary')

    summary_df_new['Error'] = summary_df_new.iloc[:, 0].str.extract(r"'([^']*)'")
    summary_df_old['Error'] = summary_df_new.iloc[:, 0].str.extract(r"'([^']*)'")
    summary_df_new = pd.merge(summary_df_new, summary_df_old[['Assigned To:','Error']], how='left',on='Error')

    load_dotenv()
    meta_data_path= os.getenv('meta_data_path')
    meta_data_df = pd.read_excel(meta_data_path)
    summary_df_new = pd.merge(summary_df_new, meta_data_df[['Description','ErrCode']], how='left',right_on='ErrCode', left_on='Error').drop(columns=['ErrCode'])

    summary_df_new['Status'] = summary_df_new['Summary'].apply(classify_status)

    summary_df_new['Error_Link'] = summary_df_new['Error'].apply(
    lambda x: f'=HYPERLINK("#{x}!A1", "{x}")'
    )

    summary_df_new = summary_df_new[['Status', 'Assigned To:', 'Error_Link', 'Summary', 'Description']]

    with pd.ExcelWriter(result_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        summary_df_new.to_excel(writer, sheet_name="Comparison Summary", index=False)

    # Load workbook
    wb = load_workbook(result_path)
    ws = wb['Comparison Summary']
    # Apply hyperlink style to the entire "Error_Link" column (assuming column C, adjust as needed)
    for cell in ws['C'][1:]:  # skip header
        cell.font = Font(color="0000FF", underline="single")
    # Save workbook
    wb.save(result_path)

def add_summary_sheet_hyperlinks(result_path):
   # Add hyperlinks to each page that links to the summary sheet
    wb = load_workbook(result_path)

    # Sheet to link to
    sheet_name = "Comparison Summary"

    # Loop through all sheets except the one we are linking to
    for ws in wb.worksheets:
        if ws.title != sheet_name:
            # Add hyperlink in cell A1 (or any cell you want)
            ws.cell(row=2, column=1).value = f'=HYPERLINK("#\'{sheet_name}\'!A1", "{sheet_name}")'
            # Optional: style as hyperlink (blue & underlined)
            from openpyxl.styles import Font
            ws.cell(row=2, column=1).font = Font(color="0000FF", underline="single")

    # Save workbook
    wb.save(result_path)

if __name__ == "__main__":

    file_path1 = select_file('first')
    file_path2 = select_file('second')
    dir = os.path.dirname(file_path1)
    os.chdir(dir)
    file_name1 = file_path1.split('/')[-1]
    file_name2 = file_path2.split('/')[-1]
    save_name = f'{file_name1[:-5]}_{file_name2[:-5]}.xlsx'
    if os.path.exists(save_name):
        os.remove(save_name) # Remove existing file if it exists to avoid overwrite issues

    change_log(file_path1, file_path2 ,save_name)
    summary_page(file_path1,save_name)
    add_summary_sheet_hyperlinks(save_name)
    print(f"Comparison complete.")

